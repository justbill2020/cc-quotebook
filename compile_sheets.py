#!/usr/bin/env python
#
#
# Requested by Drew-CodeRGV
# Created by ldartez
# 
#
#
# compile master sheet with required information
# master sheet columns: 
#   * Manufacturer
#   * Type
#   * Part Number
#   * Description
#   * List Price
#
# Aruba Mapping:
#   Manufacturer: Aruba
#   Type: Sheet Name (Access points, Switches, Central Licensing)
#   Part Number: Column A value
#   Description: Column B value
#   List Price: Column C value
#
# Cradlepoint Mapping:
#   Manufacturer: Cradlepoint
#   Type: Column B value
#   Part Number: Column D value
#   Description: Column G value
#   List Price: Column F value
#
# Fortinet Mapping:
#   Manufacturer: Fortinet
#   Type: Sheet Name (FortiGate, Wireless Products)
#   Part Number: Column B value
#   Description: Column C value
#   List Price: Column E value
#
# Meraki Mapping:
#   Manufacturer: Meraki
#   Type: Column B value
#   Part Number: Column C value
#   Description: Column D value
#   List Price: Column F value
#
# SnapAV Mapping: (Note: use filter pre-selected for POWER)
#   Manufacturer: SnapAV
#   Type: Column A value
#   Part Number: Column B value
#   Description: Column C value
#   List Price: Column J value


from config import Maps

def filter_cells(filt, x):
    '''Return true if x does not include any substrings in filt
    '''
    passing = True
    for val in x:
        for fv in filt:
            if str(fv) in str(val):
                passing = False
                break
        if not passing:
            break
    return passing

def parse_workbook_aruba(wbin):
    '''
    Aruba Mapping:
    Manufacturer: Aruba
    Type: Sheet Name (Access points, Switches, Central Licensing)
    Part Number: Column A value
    Description: Column B value
    List Price: Column C value

    Row Filters: 'Indoor Access Points', 'Mounting Brackets',
                 'Outdoor Access Points', None
    '''
    m = Maps['aruba']
    sheets = m['sheets']
    row_filters = m['row_filters']
    mftr = m['mftr']
    sku= m['sku']
    desc= m['desc']
    listPrice= m['list']
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                filtpass = filter_cells(row_filters, list(row))
                if not filtpass:
                    continue
                else:
                    row_out = (mftr, sheet.title, row[sku], row[desc], row[listPrice])
                    result.append(row_out)
    return result

def parse_workbook_cradlepoint(wbin):
    '''
    Cradlepoint Mapping:
    Manufacturer: Cradlepoint
    Type: Column B value
    Part Number: Column D value
    Description: Column G value
    List Price: Column F value
    '''
    sheets = ['USA']
    row_filters = ['Cradlepoint USA MSRP', 'Company Confidential',
            'Products']
    types = ['Routers', 'Access Points', 'LTE Adapters', 'Performance Routers',
            'Virtual Router', 'Mobile First Responder Packages',
            'Gateways', 'FIPS', 'NetCloud', 'Threat Management',
            'Internet Security', 'Feature Licenses', 'Modems',
            'SIM-in-Box', 'Antennas', 'Cradlepoint Certified',
            'Power Supplies', 'Miscellaneous', 'COR Series Routers',
            'Accessories', 'AER Series Routers', 'Home Office',
            'M2M']
    mftr = 'Cradlepoint'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                if str(row[3]) in ['None', 'Note', 'Part Number']:
                    continue
                filtpass = filter_cells(types, str(row[1]).split())
                if not filtpass:
                    cur_type = row[1]
                row_out = (mftr, cur_type, row[3], row[6], row[5])
                result.append(row_out)
    return result

def parse_workbook_fortinet(wbin):
    '''
    Fortinet Mapping:
    Manufacturer: Fortinet
    Type: Sheet Name (FortiGate, Wireless Products)
    Part Number: Column B value
    Description: Column C value
    List Price: Column E value

    '''
    sheets = ['FortiGate', 'Wireless Products']
    row_filters = ['None', 'SKU', 'RMA', 'Requires','HYPERLINK']
    mftr = 'Fortinet'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            for row in sheet.values:
                filtpass = filter_cells(row_filters, str(row[1]).split())
                if not filtpass:
                    continue
                for x in range(3,9):
                    if str(row[x]) != "None":
                        row_out = (mftr, s, row[1].replace("-DD", "-{}".format(str((x-3)*12))), row[2].replace('Hardware plus', 'Hardware plus {}year'.format(str(x-3))), row[x])
                        result.append(row_out)
    return result
    
def parse_workbook_meraki(wbin):
    '''

    Meraki Mapping:
    Manufacturer: Meraki
    Type: Column B value
    Part Number: Column C value
    Description: Column D value
    List Price: Column F value
    '''
    sheets = ['Report']
    row_filters = ['Cisco']
    mftr = 'Meraki'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            # skip first few rows
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if not filter_cells(row_filters, [row[1]]):
                    cur_type = str(row[1])
                    continue
                row_out = (mftr, cur_type, row[2], row[3], row[5])
                result.append(row_out)
    return result

def parse_workbook_snapav(wbin):
    '''
    SnapAV Mapping: (Note: use filter pre-selected for POWER)
    Manufacturer: SnapAV
    Type: Column A value
    Part Number: Column B value
    Description: Column C value
    List Price: Column J value

    '''
    sheets = ['Sheet 1']
    row_filters = ['Cisco']
    mftr = 'SnapAV'
    result = []
    for s in wbin.sheetnames:
        if s not in sheets:
            continue
        else:
            # process sheet rows
            sheet = wbin[s]
            # skip first few rows
            for row in sheet.iter_rows(values_only=True):
                if not 'Power' == str(row[0]):
                    continue
                row_out = (mftr, row[0], row[1], row[2], row[9])
                result.append(row_out)
    return result

def process_file(root, fileName):
    print("Current file: {}".format(fileName))
    if 'aruba' in basename(fileName).lower():
        msg = "Processing Aruba file: {}".format(fileName)
        print(msg)
        wb = load_workbook(os.path.join(root,fileName))
        rows = parse_workbook_aruba(wb)
        dout.extend(rows)
        wb.close()
    elif 'cradlepoint' in basename(fileName).lower():
        msg = "Processing Cradlepoint file: {}".format(fileName)
        print(msg)
        wb = load_workbook(os.path.join(root,fileName))
        rows = parse_workbook_cradlepoint(wb)
        dout.extend(rows)
        wb.close()
    elif 'fortinet' in basename(fileName).lower():
        msg = "Processing Fortinet file: {}".format(fileName)
        print(msg)
        wb = load_workbook(os.path.join(root,fileName))
        rows = parse_workbook_fortinet(wb)
        dout.extend(rows)
        wb.close()
    elif 'meraki' in basename(fileName).lower():
        msg = "Processing Meraki file: {}".format(fileName)
        print(msg)
        wb = load_workbook(os.path.join(root,fileName))
        rows = parse_workbook_meraki(wb)
        dout.extend(rows)
        wb.close()
    elif 'snapav' in basename(fileName).lower():
        msg = "Processing SnapAV file: {}".format(fileName)
        print(msg)
        wb = load_workbook(os.path.join(root,fileName))
        rows = parse_workbook_snapav(wb)
        dout.extend(rows)
        wb.close()

def easyWalk(_files_):
    try:
        number = 0
        while number < len(_files_):
            yield (str(os.path.dirname(_files_[number])), [], [_files_[number]])
            number +=1
    finally:
        return

if __name__ == "__main__":
    import argparse
    import os
    from os.path import basename
    from openpyxl import load_workbook, Workbook

    p = argparse.ArgumentParser()
    p.add_argument('-o', '--output', default='master.xlsx')
    p.add_argument('-i', '--input', default=os.getcwd())
    p.add_argument('-f', '--files', nargs='+', default=[])
    args = p.parse_args()

    fileList = args.files
    paths = args.input.strip()
    print("Current paths: {}".format(paths))

    dout = []
    walked = easyWalk(fileList)
    if len(fileList) == 0:
        walked = os.walk(paths)

    for root, dirs, files in walked:    
        print("Current root: {}".format(root))
        print("Current dirs: {}".format(dirs))
        print("Current files: {}".format(files))
        for f in files:
            process_file(root, f)

    wbout = Workbook()
    wsout = wbout.active
    
    hdr = ('Manufacturer', 'Type', 'Part Number', 'Description', 'List Price')
    wsout.append(hdr)
    if dout:
        for r in dout:
            wsout.append(r)
        wbout.save(args.output.strip())
    wbout.close()



